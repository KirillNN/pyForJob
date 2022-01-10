from openpyxl import load_workbook

f_name = 'ADC_BOM.xlsx'

wb = load_workbook(filename=f_name)
# sheet1 = wb['ADC_Test_5101']
# sheet2 = wb['ADC_Test_AD7760']
# sheet3 = wb['Merge']
sheet1 = wb[wb.sheetnames[0]]
sheet2 = wb[wb.sheetnames[1]]
sheet3 = wb[wb.sheetnames[2]]

# print(sheet['A2'].value)
dvs = []
counter = 1
# f = open('Output.txt', 'w', encoding='utf-8')
# data_sheet1 = {}
# data_sheet2 = {}

all_data = {}

for cell_obj in sheet1['A2':'E10000']:
    if cell_obj[0].value is None:  # если дошли до пустой ячейки
        break  # прерываем выполение цикла
    # add all_data['ADP7158ACPZ-1.8-R7'] = 1[]
    all_data[cell_obj[0].value] = [cell_obj[1].value, cell_obj[2].value, cell_obj[3].value]

# for key, value in all_data.items():
#     print(key, '->', value)

for cell_obj in sheet2['A2':'E10000']:
    if cell_obj[0].value is None:  # если дошли до пустой ячейки
        break  # прерываем выполение цикла
    if all_data.get(cell_obj[0].value) is None:  # если в all_data нет такого элемента
        all_data[cell_obj[0].value] = [cell_obj[1].value, cell_obj[2].value, cell_obj[3].value]  # добаляем
    else:  # если есть, прибавляем количество
        all_data[cell_obj[0].value] = [all_data[cell_obj[0].value][0] + cell_obj[1].value, cell_obj[2].value,
                                       cell_obj[3].value]
    # print(all_data.get(cell_obj[0].value))

for key, value in all_data.items():
    print(key, '->', value)

# for key in all_data:
#     print(key, '->', all_data[key])
#
# for item in all_data.items():
#     print(item)
sheet3['A1'] = 'Part Number'
sheet3['B1'] = 'Quantity'
sheet3['C1'] = 'Description'
sheet3['D1'] = 'Comment'
x = 2
for key, value in all_data.items():  # записываем обобщенные данные
    # print(key, '->', value)
    sheet3['A' + str(x)] = key
    sheet3['B' + str(x)] = value[0]
    sheet3['C' + str(x)] = value[1]
    sheet3['D' + str(x)] = value[2]
    x += 1

wb.save(filename=f_name)

# print(data_sheet1)
# print(data_sheet2)
# https://openpyxl.readthedocs.io/en/stable/usage.html#write-a-workbook
# sheet3[0][0] = 500

# sheet3['A1'] = 200
# wb.save(filename=f_name)

# type_nominal = cell_obj[0].value
#     if isinstance(cell_obj[0].value, int) and cell_obj[0].value == counter:
#         counter += 1
#         for cell in cell_obj:
#             if cell.value is not None:
#                 f.write(str(cell.value) + '\t')
#         f.write('\n')
# f.close()
